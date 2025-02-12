import openpyxl
# Cargar un libro de Excel
libro =openpyxl.load_workbook("C:/Users/artha/OneDrive/Escritorio/AKKODIS/python-rda/05_ejemplos/distritos-ejemplo.xlsx")
libro.sheetnames
# ['Distrito_población']
# Crear una nueva pestaña
pestaña = libro['Distrito_población'] 
pestaña['C1']= 'Ranking'
# Guardar el libro con la nueva pestaña
pestaña2 = libro.create_sheet(title='Nueva pestaña')
libro.save("C:/Users/artha/OneDrive/Escritorio/AKKODIS/python-rda/05_ejemplos/distritos-ejemplo_v2.xlsx")
