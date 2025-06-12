import openpyxl #pip install openpyxl
# Cargar un libro de Excel
libro =openpyxl.load_workbook("C:/Users/artha/OneDrive/Escritorio/AKKODIS/python-rda/05_xlsx_docx_pdf/distritos-ejemplo.xlsx")
libro.sheetnames
# ['Distrito_población']
pestaña = libro['Distrito_población']
pestaña['A1'].value
# 'DISTRITO'
pestaña.cell(row=2, column=1).value  
# 'Arganzuela'
for i in range(1,6):
    print(pestaña.cell(row=i, column=1).value)
# DISTRITO
# Arganzuela
# Centro
# Chamberí
# Moncloa-Aravaca